#!/usr/bin/env python3
"""
Télécharge et vérifie les timestamps dans l'Excel créé
"""
import sys
from pathlib import Path
import openpyxl

sys.path.insert(0, str(Path(__file__).parent / 'src'))

from drive_manager import DriveManager

# ID de l'Excel créé (depuis les logs)
EXCEL_ID = "1Bo6c06DLfQbiAk60lw-z3MEEE92dH-um"

def main():
    manager = DriveManager(credentials_path='./config/credentials.json')

    # Télécharger l'Excel
    file_name = 'check_excel.xlsx'
    full_path = '/tmp/check_excel.xlsx'
    print(f"📥 Téléchargement de l'Excel...")

    manager.download_file(EXCEL_ID, file_name, full_path)
    print(f"✅ Téléchargé: {full_path}")
    print("")

    # Lire l'Excel
    wb = openpyxl.load_workbook(full_path)
    ws = wb.active

    print("=" * 80)
    print("📊 TIMESTAMPS DANS L'EXCEL")
    print("=" * 80)
    print("")

    # En-têtes
    headers = [cell.value for cell in ws[1]]
    print(f"Colonnes: {headers}")
    print("")

    # Données (colonnes: Numéro, Groupe, Sous-segment, Total, Début(s), Fin(s), ...)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        numero = row[0]
        groupe = row[1]  # S1, S3, S4, S5
        sous_segment = row[2]
        start_seconds = row[4]  # Début (secondes)
        end_seconds = row[5]    # Fin (secondes)
        start_time = row[6]     # Début (HH:MM:SS)
        end_time = row[7]       # Fin (HH:MM:SS)
        duration = row[8]

        print(f"Row {row_idx}: {groupe} (sous-segment {sous_segment})")
        print(f"  Start:  {start_seconds}s ({start_time})")
        print(f"  End:    {end_seconds}s ({end_time})")
        print(f"  Durée:  {duration}s")
        print("")

if __name__ == '__main__':
    main()
